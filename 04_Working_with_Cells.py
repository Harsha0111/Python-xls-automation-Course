# Open or create a workbook
from openpyxl.styles import Font, Alignment
from openpyxl import Workbook

wb = Workbook()
ws = wb.active

# Reading and writing data to cells
ws["A1"] = "Hello World"
# Formatting Cells
ws["A1"].font = Font(bold=True, color="FF5733")
ws["A1"].alignment = Alignment(horizontal="center")

# Merging and splitting cells
ws.merge_cells("A1:C1")
# ws.unmerge_cells("A1:C1")

# Using Formulas
ws["A3"] = 10
ws["B3"] = 20
ws["C3"] = "=A3+B3"

# Save changes
wb.save("output/04_Working_with_Cells.xlsx")
print("Cell operations completed and saved as '04_Working_with_Cells.xlsx'")
